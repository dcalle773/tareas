import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image

def cargar_archivo(ruta):
    """
    Carga el archivo especificado y lo convierte en un DataFrame.

    :param ruta: Ruta del archivo.
    :return: DataFrame con los datos del archivo.
    """
    if ruta.endswith('.csv'):
        df = pd.read_csv(ruta)
    elif ruta.endswith('.xlsx'):
        df = pd.read_excel(ruta, sheet_name='Sheet1')  # Cambiar 'Sheet1' si es necesario
    else:
        raise ValueError("El archivo debe ser .csv o .xlsx")
    
    return df

def generar_graficos(df):
    """
    Genera gráficos de barras, torta, histograma, regresión lineal, Random Forest y análisis de correlación.

    :param df: DataFrame con los datos.
    """
    # Gráfico de barras
    plt.figure(figsize=(10, 6))
    df['Club'].value_counts().plot(kind='bar')
    plt.title('Distribución por Club')
    plt.xlabel('Club')
    plt.ylabel('Frecuencia')
    plt.tight_layout()
    plt.savefig('grafico_barras.png')
    
    # Gráfico de torta
    plt.figure(figsize=(8, 8))
    df['Nationality'].value_counts().plot(kind='pie', autopct='%1.1f%%')
    plt.title('Distribución por Nacionalidad')
    plt.ylabel('')
    plt.tight_layout()
    plt.savefig('grafico_torta.png')
    
    # Histograma
    plt.figure(figsize=(10, 6))
    df['Age'].hist(bins=10)
    plt.title('Distribución de Edad')
    plt.xlabel('Edad')
    plt.ylabel('Frecuencia')
    plt.tight_layout()
    plt.savefig('histograma.png')
    
    # Gráfico de correlación entre edad y cantidad de goles
    plt.figure(figsize=(10, 6))
    sns.regplot(x='Age', y='Goals', data=df)
    plt.title('Correlación entre Edad y Cantidad de Goles')
    plt.xlabel('Edad')
    plt.ylabel('Goles')
    plt.tight_layout()
    plt.savefig('correlacion_edad_goles.png')
    
    # Gráfico de goles por posición
    plt.figure(figsize=(10, 6))
    df.groupby('Position')['Goals'].sum().plot(kind='bar')
    plt.title('Goles por Posición')
    plt.xlabel('Posición')
    plt.ylabel('Goles')
    plt.tight_layout()
    plt.savefig('goles_por_posicion.png')
    
    # Gráfico de distribución de la precisión de los disparos
    plt.figure(figsize=(10, 6))
    df['Shooting accuracy %'].dropna().hist(bins=10)
    plt.title('Distribución de la Precisión de los Disparos')
    plt.xlabel('Precisión de Disparos (%)')
    plt.ylabel('Frecuencia')
    plt.tight_layout()
    plt.savefig('precision_disparos.png')
    
    # Gráfico de distribución de asistencias por club
    plt.figure(figsize=(10, 6))
    df.groupby('Club')['Assists'].sum().plot(kind='bar')
    plt.title('Distribución de Asistencias por Club')
    plt.xlabel('Club')
    plt.ylabel('Asistencias')
    plt.tight_layout()
    plt.savefig('asistencias_por_club.png')
    
    # Gráfico de comparación de goles y asistencias por posición
    plt.figure(figsize=(10, 6))
    df_grouped = df.groupby('Position').agg({'Goals': 'sum', 'Assists': 'sum'})
    df_grouped.plot(kind='bar')
    plt.title('Comparación de Goles y Asistencias por Posición')
    plt.xlabel('Posición')
    plt.ylabel('Cantidad')
    plt.tight_layout()
    plt.savefig('goles_asistencias_posicion.png')

    # Verificar si las columnas 'Height' y 'Weight' existen
    if 'Height' in df.columns and 'Weight' in df.columns:
        # Regresión Lineal
        df = df.dropna(subset=['Height', 'Weight'])
        X = df[['Height']]
        y = df['Weight']
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)
        model = LinearRegression()
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        
        # Predicciones a futuro
        future_heights = np.linspace(df['Height'].min(), df['Height'].max() + 10, 100).reshape(-1, 1)
        future_weights = model.predict(future_heights)
        
        plt.figure(figsize=(10, 6))
        plt.scatter(X_test, y_test, color='red', label='Datos Reales')
        plt.plot(X_test, y_pred, color='blue', linewidth=2, label='Regresión Lineal')
        plt.plot(future_heights, future_weights, color='green', linestyle='--', label='Predicciones Futuras')
        plt.title('Regresión Lineal entre Altura y Peso')
        plt.xlabel('Altura')
        plt.ylabel('Peso')
        plt.legend()
        plt.tight_layout()
        plt.savefig('regresion_lineal.png')
        
        # Random Forest
        model_rf = RandomForestRegressor(n_estimators=100, random_state=0)
        model_rf.fit(X_train, y_train)
        y_pred_rf = model_rf.predict(X_test)
        
        # Predicciones a futuro
        future_weights_rf = model_rf.predict(future_heights)
        
        plt.figure(figsize=(10, 6))
        plt.scatter(X_test, y_test, color='red', label='Datos Reales')
        plt.scatter(X_test, y_pred_rf, color='green', label='Predicciones Random Forest')
        plt.plot(future_heights, future_weights_rf, color='blue', linestyle='--', label='Predicciones Futuras Random Forest')
        plt.title('Predicciones Random Forest')
        plt.xlabel('Altura')
        plt.ylabel('Peso')
        plt.legend()
        plt.tight_layout()
        plt.savefig('random_forest.png')
    
    # Insertar gráficos en un archivo Excel
    wb = load_workbook('Out.xlsx') if os.path.exists('Out.xlsx') else Workbook()
    ws = wb.create_sheet('Gráficos')

    for i, img_path in enumerate([
        'grafico_barras.png', 'grafico_torta.png', 'histograma.png', 
        'correlacion_edad_goles.png', 'goles_por_posicion.png', 
        'precision_disparos.png', 'asistencias_por_club.png', 
        'goles_asistencias_posicion.png'
    ]):
        img = Image(img_path)
        ws.add_image(img, f'A{i*20+1}')
    
    if 'Height' in df.columns and 'Weight' in df.columns:
        for i, img_path in enumerate(['regresion_lineal.png', 'random_forest.png'], start=8):
            img = Image(img_path)
            ws.add_image(img, f'A{i*20+1}')
    
    wb.save('Out.xlsx')

def procesar_datos():
    """
    Función principal que maneja el flujo del proceso ETL.
    """
    ruta = filedialog.askopenfilename(filetypes=[("Archivos CSV y Excel", "*.csv *.xlsx")])
    if not ruta or not os.path.isfile(ruta):
        messagebox.showerror("Error", "Por favor, seleccione un archivo válido.")
        return

    df = cargar_archivo(ruta)
    
    if df.empty:
        messagebox.showinfo("Información", "No se encontraron datos para procesar.")
    else:
        # Exportar el DataFrame a Excel
        try:
            df.to_excel('Out.xlsx', index=False)
            generar_graficos(df)
            messagebox.showinfo("Éxito", "Datos exportados a Out.xlsx y gráficos generados.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar datos o generar gráficos: {e}")

        # Mostrar el DataFrame en una nueva ventana
        mostrar_dataframe(df)

def mostrar_dataframe(df):
    """
    Muestra el DataFrame en una nueva ventana.

    :param df: DataFrame a mostrar.
    """
    ventana_df = tk.Toplevel(ventana)
    ventana_df.title("Dataset Final")

    # Crear un widget Text para mostrar el DataFrame
    text_widget = tk.Text(ventana_df)
    text_widget.insert(tk.END, df.to_string())
    text_widget.pack(fill=tk.BOTH, expand=True)

# Configuración de la interfaz gráfica
ventana = tk.Tk()
ventana.title("Proceso ETL - Archivos")
ventana.geometry("600x250")

boton_procesar = tk.Button(ventana, text="Seleccionar y Procesar Archivo", command=procesar_datos)
boton_procesar.pack(pady=10)

ventana.mainloop()
